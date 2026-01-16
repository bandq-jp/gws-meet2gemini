from __future__ import annotations

from dataclasses import dataclass
import html
import io
import logging
import re
import zipfile
import xml.etree.ElementTree as ET
from typing import Optional, Dict, Any, Iterable, List, Tuple

try:  # optional dependency
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - fallback path
    load_workbook = None


_META_DATETIME_RE = re.compile(
    r"(?P<y>20\d{2})[./-](?P<m>\d{1,2})[./-](?P<d>\d{1,2})\s*(?P<h>\d{1,2})[:_](?P<min>\d{2})"
)
_FILENAME_DATETIME_RE = re.compile(
    r"(?P<y>20\d{2})[./-](?P<m>\d{1,2})[./-](?P<d>\d{1,2})[ _-]+(?P<h>\d{1,2})[:_](?P<min>\d{2})"
)
_DURATION_RE = re.compile(r"[·・]\s*(?P<mins>\d{1,4})\s*mins?", re.IGNORECASE)


@dataclass
class NottaParseResult:
    title: Optional[str]
    meeting_datetime: Optional[str]
    duration_mins: Optional[int]
    text_content: str
    speaker_stats: Dict[str, int]
    meta_line: Optional[str]
    sheet_name: Optional[str]
    row_count: int


class NottaXlsxParser:
    """Parse Notta Excel transcripts into normalized text content."""

    def __init__(self) -> None:
        self.logger = logging.getLogger(__name__)

    def parse(self, data: bytes, filename: Optional[str] = None) -> NottaParseResult:
        if load_workbook is not None:
            try:
                return self._parse_with_openpyxl(data, filename)
            except Exception as e:  # pragma: no cover - fallback only when openpyxl fails
                self.logger.warning("openpyxl parse failed, falling back to XML parser: %s", e)
        return self._parse_with_xml(data, filename)

    def _parse_with_openpyxl(self, data: bytes, filename: Optional[str]) -> NottaParseResult:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        best = None
        for sheet in wb.worksheets:
            rows = self._collect_rows_openpyxl(sheet)
            score = sum(1 for s, t in rows if s or t)
            if best is None or score > best[0]:
                best = (score, sheet.title, rows, sheet)

        if best is None:
            return self._empty_result(filename)

        _, sheet_name, rows, sheet = best
        title = self._stringify(sheet["A1"].value)
        meta_line = self._stringify(sheet["A2"].value)
        return self._build_result(rows, title, meta_line, filename, sheet_name)

    def _collect_rows_openpyxl(self, sheet) -> List[Tuple[Optional[str], Optional[str]]]:
        rows: List[Tuple[Optional[str], Optional[str]]] = []
        for row in sheet.iter_rows(min_row=4, values_only=True):
            speaker = self._stringify(row[0]) if len(row) > 0 else None
            text = self._stringify(row[1]) if len(row) > 1 else None
            rows.append((speaker, text))
        return rows

    def _parse_with_xml(self, data: bytes, filename: Optional[str]) -> NottaParseResult:
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as zf:
                shared_strings = self._load_shared_strings(zf)
                sheets = [n for n in zf.namelist() if n.startswith("xl/worksheets/sheet")]
                if not sheets:
                    return self._empty_result(filename)

                best_rows: List[Tuple[Optional[str], Optional[str]]] = []
                best_sheet = None
                best_score = -1
                best_title = None
                best_meta = None
                for sheet_name in sheets:
                    xml = zf.read(sheet_name).decode("utf-8")
                    title, meta_line, rows = self._parse_sheet_xml(xml, shared_strings)
                    score = sum(1 for s, t in rows if s or t)
                    if score > best_score:
                        best_score = score
                        best_rows = rows
                        best_sheet = sheet_name
                        best_title = title
                        best_meta = meta_line

                return self._build_result(best_rows, best_title, best_meta, filename, best_sheet)
        except Exception as e:
            self.logger.error("Failed to parse Notta xlsx via XML: %s", e)
            return self._empty_result(filename)

    def _load_shared_strings(self, zf: zipfile.ZipFile) -> List[str]:
        if "xl/sharedStrings.xml" not in zf.namelist():
            return []
        xml = zf.read("xl/sharedStrings.xml").decode("utf-8")
        root = ET.fromstring(xml)
        ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        strings = []
        for si in root.findall("a:si", ns):
            texts = []
            for t in si.findall(".//a:t", ns):
                if t.text is not None:
                    texts.append(t.text)
            strings.append(html.unescape("".join(texts)))
        return strings

    def _parse_sheet_xml(
        self, xml: str, shared_strings: List[str]
    ) -> Tuple[Optional[str], Optional[str], List[Tuple[Optional[str], Optional[str]]]]:
        ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        root = ET.fromstring(xml)
        title = None
        meta_line = None
        rows: List[Tuple[Optional[str], Optional[str]]] = []

        for row in root.findall(".//a:sheetData/a:row", ns):
            idx = int(row.get("r") or 0)
            cells: Dict[str, Optional[str]] = {}
            for c in row.findall("a:c", ns):
                ref = c.get("r") or ""
                col = "".join(ch for ch in ref if ch.isalpha())
                value = self._read_cell_value(c, shared_strings, ns)
                if col:
                    cells[col] = value
            if idx == 1:
                title = cells.get("A")
            elif idx == 2:
                meta_line = cells.get("A")
            elif idx >= 4:
                rows.append((cells.get("A"), cells.get("B")))

        return title, meta_line, rows

    def _read_cell_value(
        self, cell: ET.Element, shared_strings: List[str], ns: Dict[str, str]
    ) -> Optional[str]:
        t = cell.get("t")
        if t == "inlineStr":
            text_el = cell.find("a:is/a:t", ns)
            return html.unescape(text_el.text) if text_el is not None and text_el.text else None
        v_el = cell.find("a:v", ns)
        if v_el is None or v_el.text is None:
            return None
        raw = v_el.text
        if t == "s":
            try:
                idx = int(raw)
                return shared_strings[idx] if idx < len(shared_strings) else raw
            except Exception:
                return raw
        return raw

    def _build_result(
        self,
        rows: Iterable[Tuple[Optional[str], Optional[str]]],
        title: Optional[str],
        meta_line: Optional[str],
        filename: Optional[str],
        sheet_name: Optional[str],
    ) -> NottaParseResult:
        meeting_dt = self._parse_datetime(meta_line, filename)
        duration = self._parse_duration(meta_line)
        text_content, stats, row_count = self._build_text(rows)
        return NottaParseResult(
            title=title or None,
            meeting_datetime=meeting_dt,
            duration_mins=duration,
            text_content=text_content,
            speaker_stats=stats,
            meta_line=meta_line,
            sheet_name=sheet_name,
            row_count=row_count,
        )

    def _build_text(
        self, rows: Iterable[Tuple[Optional[str], Optional[str]]]
    ) -> Tuple[str, Dict[str, int], int]:
        lines: List[str] = []
        speaker_stats: Dict[str, int] = {}
        current_speaker: Optional[str] = None
        current_text: List[str] = []
        row_count = 0

        def flush():
            nonlocal current_speaker, current_text
            if current_speaker and current_text:
                text = " ".join(t for t in current_text if t)
                if text:
                    lines.append(f"{current_speaker}: {text}")
            current_speaker = None
            current_text = []

        for speaker_raw, text_raw in rows:
            speaker = self._stringify(speaker_raw)
            text = self._stringify(text_raw)
            if not speaker and not text:
                continue
            row_count += 1
            if speaker:
                speaker_stats[speaker] = speaker_stats.get(speaker, 0) + 1
            if speaker and speaker != current_speaker:
                flush()
                current_speaker = speaker
            if text:
                if not current_speaker:
                    # No speaker; attach to previous if possible, otherwise mark unknown
                    current_speaker = current_speaker or "UNKNOWN"
                current_text.append(text)

        flush()
        return "\n".join(lines), speaker_stats, row_count

    def _parse_datetime(self, meta_line: Optional[str], filename: Optional[str]) -> Optional[str]:
        for source in (meta_line, filename):
            if not source:
                continue
            m = _META_DATETIME_RE.search(str(source)) or _FILENAME_DATETIME_RE.search(str(source))
            if m:
                y = int(m.group("y"))
                mo = int(m.group("m"))
                d = int(m.group("d"))
                h = int(m.group("h"))
                mi = int(m.group("min"))
                return f"{y:04d}/{mo:02d}/{d:02d} {h:02d}:{mi:02d}"
        return None

    def _parse_duration(self, meta_line: Optional[str]) -> Optional[int]:
        if not meta_line:
            return None
        m = _DURATION_RE.search(str(meta_line))
        if not m:
            return None
        try:
            return int(m.group("mins"))
        except Exception:
            return None

    def _stringify(self, value: Any) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        return text if text else None

    def _empty_result(self, filename: Optional[str]) -> NottaParseResult:
        return NottaParseResult(
            title=None,
            meeting_datetime=self._parse_datetime(None, filename),
            duration_mins=None,
            text_content="",
            speaker_stats={},
            meta_line=None,
            sheet_name=None,
            row_count=0,
        )
